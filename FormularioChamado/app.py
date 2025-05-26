from flask import Flask, render_template, request, redirect, url_for, send_file
from openpyxl import load_workbook, Workbook
from datetime import datetime
import io
import os

app = Flask(__name__)
app.secret_key = 'sua_chave_secreta_aqui'

ARQUIVO_EXCEL = r"./Downloads/Chamados_CEMIG.xlsx"
os.makedirs(os.path.dirname(ARQUIVO_EXCEL), exist_ok=True)

# Cria planilha se não existir
if not os.path.exists(ARQUIVO_EXCEL):
    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados"
    ws.append([
        "data_hora", "chamado", "localicade", "n_medidor",
        "uc_instalacao", "medidor_atual", "descricao", "status"
    ])
    wb.save(ARQUIVO_EXCEL)

@app.route('/', methods=["GET", "POST"])
def index():
    if request.method == "POST":
        chamado = request.form.get("chamado")
        horario = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        localicade = request.form.get("localicade")
        n_medidor = request.form.get("n_medidor")
        uc_instalacao = request.form.get("uc_instalacao")
        medidor_atual = request.form.get("medidor_atual")
        descricao = request.form.get("descricao")
        status = request.form.get("status")

        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
        ws.append([
            horario, chamado, localicade, n_medidor,
            uc_instalacao, medidor_atual, descricao, status
        ])
        wb.save(ARQUIVO_EXCEL)

        return redirect(url_for('index', success=1))

    success = request.args.get('success')
    localidades, medidores, ucs = set(), set(), set()

    if os.path.exists(ARQUIVO_EXCEL):
        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2]: localidades.add(row[2])
            if row[3]: medidores.add(row[3])
            if row[4]: ucs.add(row[4])

    return render_template(
        'form.html',
        success=success,
        localidades=sorted(localidades),
        medidores=sorted(medidores),
        ucs=sorted(ucs, key=str),
        now=datetime.now()
    )

@app.route('/lista_chamados')
def lista_chamados():
    chamados = []
    if os.path.exists(ARQUIVO_EXCEL):
        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            chamados.append({
                "data_hora": row[0],
                "chamado": row[1],
                "localicade": row[2],
                "n_medidor": row[3],
                "uc_instalacao": row[4],
                "medidor_atual": row[5],
                "descricao": row[6],
                "status": row[7]
            })

    return render_template('lista_chamados.html', chamados=chamados)

from flask import send_file
import io

@app.route('/exportar_todos')
def exportar_todos():
    # Simplesmente retorna o arquivo Excel salvo
    if os.path.exists(ARQUIVO_EXCEL):
        return send_file(ARQUIVO_EXCEL, as_attachment=True)
    else:
        return "Arquivo não encontrado", 404


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
