from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = 'sua_chave_secreta_aqui'

ARQUIVO_EXCEL = r".\Downloads\Chamados_CEMIG.xlsx"
os.makedirs(os.path.dirname(ARQUIVO_EXCEL), exist_ok=True)

if not os.path.exists(ARQUIVO_EXCEL):
    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados"
    ws.append(["data_hora", "chamado", "Localicade", "Estabelecimento", "N° Medidor", "UC/Instalação", "Medidor atual", "Descrição"])
    wb.save(ARQUIVO_EXCEL)

@app.route('/', methods=["GET", "POST"])
def index():
    if request.method == "POST":
        chamado = request.form.get("chamado")
        horario = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        localicade = request.form.get("localicade")
        estabelecimento = request.form.get("estabelecimento")
        n_medidor = request.form.get("n_medidor")
        uc_instalacao = request.form.get("uc_instalacao")
        medidor_atual = request.form.get("medidor_atual")
        descricao = request.form.get("descricao")
        

        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
        ws.append([horario, chamado, localicade, estabelecimento, n_medidor, uc_instalacao, medidor_atual, descricao])
        wb.save(ARQUIVO_EXCEL)

        return redirect(url_for('index', success=1))


    success = request.args.get('success')

    chamados = []
    chamados_unicos = []

    if os.path.exists(ARQUIVO_EXCEL):
        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row) + [""] * (8 - len(row))
            
            chamados.append({
                "data_hora": row[0],
                "chamado": row[1],
                "localicade": row[2],
                "estabelecimento": row[3],
                "n_medidor": row[4],
                "uc_instalacao": row[5],
                "medidor_atual": row[6],
                "descricao": row[7]
            })
        chamados_unicos = sorted(set([c["chamado"] for c in chamados if c["chamado"] is not None]))


    return render_template(
        'form.html',
        success=success,
        chamados=chamados,
        chamados_unicos=chamados_unicos,
        now=datetime.now()
    )

@app.route('/registrar', methods=['POST'])
def registrar():
    chamado = request.form['chamado']
    datahora = datetime.now().strftime("%d/%m/%Y %H:%M")
    localidade = request.form['localidade']
    estabelecimento = request.form['estabelecimento']
    n_medidor = request.form['n_medidor']
    uc = request.form['uc']
    medidor_atual = request.form['medidor_atual']
    descricao = request.form['descricao']

    wb = load_workbook(ARQUIVO_EXCEL)
    ws = wb.active
    ws.append([
    chamado, datahora, localidade, estabelecimento,
    n_medidor, uc, medidor_atual, descricao
    ])

    wb.save(ARQUIVO_EXCEL)

    return redirect(url_for('index'))

@app.route('/listar')
def lista_chamados():
    chamados = []
    if os.path.exists(ARQUIVO_EXCEL):
        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row) + [""] * (8 - len(row))  # evita erros se faltar alguma coluna
            chamados.append({
                "data_hora": row[0],
                "chamado": row[1],
                "localidade": row[2],
                "estabelecimento": row[3],
                "n_medidor": row[4],
                "uc_instalacao": row[5],
                "medidor_atual": row[6],
                "descricao": row[7]
            })
    return render_template('lista_chamados.html', chamados=chamados)

@app.route('/exportar_todos')
def exportar_todos():
    # Lógica de exportação aqui (ex: gerar CSV, Excel, etc)
    return 'Função de exportar todos os chamados ainda não implementada'


@app.route('/excluir/<chamado>')
def excluir_chamado(chamado):
    wb = load_workbook(ARQUIVO_EXCEL)
    ws = wb.active
    ws.delete_rows(chamado + 2)
    wb.save(ARQUIVO_EXCEL)
    return redirect(url_for('lista_chamados'))

@app.route('/editar/<chamado>', methods=['GET', 'POST'])
def edita_chamado(chamado):
    wb = load_workbook(ARQUIVO_EXCEL)
    ws = wb.active

    linhas = list(ws.iter_rows(min_row=2, values_only=False))

    # Procurar o índice da linha com o chamado igual ao recebido na URL
    indice = None
    for i, linha in enumerate(linhas):
        valor_chamado = linha[1].value
        if valor_chamado == chamado:
            indice = i
            break

    if indice is None:
        return f"Chamado '{chamado}' não encontrado.", 404
    
    if request.method == 'POST':
        linha = linhas[indice]
        linha[1].value = request.form['chamado']
        linha[2].value = request.form['localidade']
        linha[3].value = request.form['estabelecimento']
        linha[4].value = request.form['n_medidor']
        linha[5].value = request.form['uc']
        linha[6].value = request.form['medidor_atual']
        linha[7].value = request.form['descricao']

        wb.save(ARQUIVO_EXCEL)
        return redirect(url_for('lista_chamados'))

    dados = [cell.value for cell in linhas[indice]]
    return render_template('edita_chamado.html', chamado=dados)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
